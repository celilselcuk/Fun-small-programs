import random
from openpyxl import load_workbook
from difflib import SequenceMatcher

workbook = load_workbook("words.xlsx")
sheet = workbook.active

catch = [",", " ", "("]

def make_word_list():
    qa_pairs = []
    for row in sheet.iter_rows(min_row=1, max_col=2):
        if row[0].value and row[1].value:
            qa_pairs.append((row[0].value, row[1].value))

    for pos in range(len(qa_pairs)):
        word = qa_pairs[pos][0]
        for i in range(len(word)):
            if word[i] in catch:
                qa_pairs[pos] = (word[0:i], qa_pairs[pos][1])
                break
    return qa_pairs

qa_pairs = make_word_list()
random.shuffle(qa_pairs)

def line_separation():
    print("")
    print("-" * 50) 
    print("")

def check_answer(user_answer, correct_answer):
    possible_answers = correct_answer.split('/')
    for i in range(len(possible_answers)):
        possible_answers[i] = possible_answers[i].strip().lower()
    
    user_answers = user_answer.split(',')
    for i in range(len(user_answers)):
        user_answers[i] = user_answers[i].strip().lower()
        
    feedback = []
    matched = []
    for ua in user_answers:
        best_match = possible_answers[0]
        best_ratio = SequenceMatcher(None, ua, best_match).ratio()
        for pa in possible_answers:
            match_ratio = SequenceMatcher(None, ua, pa).ratio()
            if match_ratio > best_ratio:
                best_match = pa
                best_ratio = match_ratio
        if best_ratio == 1.0:
            matched.append(best_match)
        else:
            feedback.append(f"'{ua}' is incorrect or incomplete. Best match is '{best_match}'")

    missing_answers = []
    for pa in possible_answers:
        if pa not in matched:
            missing_answers.append(pa)
    if missing_answers:
        feedback.append(f"Missing: {', '.join(missing_answers)}")

    return len(feedback) == 0, ', '.join(feedback)

def ask_questions():
    score = 0
    total_questions = len(qa_pairs)
    q_number = 1
    wrong = []
    for i in range(total_questions):
        word, definition = qa_pairs[i]
        print(f"{q_number}) What is the definition of: {word}")
        answer = input("Your answer (separate multiple answers with commas): ")

        correct, feedback = check_answer(answer, definition)
        
        if correct:
            print("Correct!")
            score += 1
        else:
            print(f"Incorrect --> {feedback}")
            print(f"The correct definition is: {definition}")
            wrong.append(word)
            
        line_separation()
        
        q_number += 1
        
    percentage = (score / total_questions) * 100
    print(f"Your final score: {score}/{total_questions}")
    print(f"Percentage: {percentage:.2f}%")
    if len(wrong) > 0:
        print("You got the following words wrong: ")
        line_separation()
        for incorrect_word in wrong:
            print(incorrect_word)
        line_separation()
    else:
        print("Well done: you got nothing wrong!")

ask_questions()
input()
